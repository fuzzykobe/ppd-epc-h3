"""
Ingest ONS small-area household income estimates → data/staged/stg_income.parquet

Actual files found in data/raw/income/ (inspected 2025-05-02):

  filename                                                   FY ending  Start yr  Unit    MSOA boundary
  1smallareaincomeestimatesdatatcm77420299.xls               2012       2011      weekly  2011
  1smallareaincomeestimatesdataupdate.xls                    2014       2013      weekly  2011
  1smallareaincomeestimatesdata.xls                          2016       2015      annual  2011
  incomeestimatesforsmallareasdatasetfinancialyearending20181.xls  2018  2017    annual  2011
  saiefy1920finalqaddownload280923.xlsx                      2020       2019      annual  2011
  datasetfinal_2023.xlsx                                     2023       2022      annual  2021

Year convention: financial_year stored as START of financial year (FY ending 2020 → stored 2019),
matching the convention in 13_build_price_income_mart.py (transfer_year=2019 joins to FY2019/20).

MSOA boundary: 2011-boundary files are remapped to 2021 codes via the ONS best-fit crosswalk
at data/raw/income/msoa_2011_to_2021_lookup.csv (1-to-1 for 7080 unchanged codes, 121 remapped).

For changed 2011→2021 codes the 2021 code is substituted in place.
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any

import click
import duckdb
from loguru import logger

from pipeline.config import DATA_RAW, DATA_STAGED, DUCKDB_MEMORY, DUCKDB_THREADS

INCOME_DIR    = DATA_RAW / "income"
CROSSWALK_CSV = INCOME_DIR / "msoa_2011_to_2021_lookup.csv"
ONS_PARQUET   = DATA_STAGED / "stg_ons.parquet"

# ─── Sheet-selection priority ────────────────────────────────────────────────
# We want net disposable annual income. These are tried in order.
PREFERRED_SHEETS = [
    "Net annual income",
    "Net income before housing costs",
    "Net weekly income",          # older files (2011/12, 2013/14): must × 52
    "Total annual income",        # fallback
    "Total weekly income",        # last resort
]
WEEKLY_SHEETS = {"Net weekly income", "Total weekly income",
                 "Net income before housing costs"}

# Older files publish equivalised BHC as the third sheet; the unequivalised
# "Net annual income" / "Net weekly income" is preferred.


# ─── Helpers ────────────────────────────────────────────────────────────────

def _load_crosswalk() -> dict[str, str]:
    """Returns {msoa11cd: msoa21cd}. Only entries where codes differ are stored."""
    if not CROSSWALK_CSV.exists():
        logger.warning("MSOA crosswalk not found at {}; 2011 codes kept as-is", CROSSWALK_CSV)
        return {}
    mapping: dict[str, str] = {}
    with open(CROSSWALK_CSV, encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            c11, c21 = row["MSOA11CD"].strip(), row["MSOA21CD"].strip()
            if c11 != c21:
                mapping[c11] = c21
    logger.info("Crosswalk loaded: {} 2011→2021 remaps", len(mapping))
    return mapping


def _year_from_text(text: str) -> int | None:
    """Extract FY-ending year from text like 'financial year ending 2020' or '2015/16'."""
    # "financial year ending YYYY"
    m = re.search(r"financial\s+year\s+ending\s+(?:march\s+)?(\d{4})", text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    # "YYYY/YY" pattern (e.g. "2015/16")
    m = re.search(r"\b(20\d{2})/(\d{2})\b", text)
    if m:
        return int(m.group(1)) + 1  # "2015/16" → ending year 2016
    return None


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", "").replace("£", "").strip())
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _extract_sheet_xls(ws: Any) -> tuple[list[dict], bool]:
    """
    Parse an xlrd worksheet. Returns (records, sheet_is_weekly).
    income is the raw value from the sheet — caller applies ×52 if needed.
    """
    sheet_is_weekly = "weekly" in ws.name.lower()

    # Find header row (first row where col-0 cell text == "MSOA code")
    header_row = None
    for i in range(min(20, ws.nrows)):
        if str(ws.cell_value(i, 0)).strip().lower() == "msoa code":
            header_row = i
            break
    if header_row is None:
        return [], sheet_is_weekly

    # Income value is always column 6 in every ONS file we have
    records = []
    for i in range(header_row + 1, ws.nrows):
        msoa = str(ws.cell_value(i, 0)).strip()
        if not re.match(r"^[EW]\d{8}$", msoa):
            continue
        income = _to_float(ws.cell_value(i, 6))
        if income is None:
            continue
        records.append({
            "msoa_code": msoa,
            "msoa_name": str(ws.cell_value(i, 1)).strip(),
            "la_code":   str(ws.cell_value(i, 2)).strip(),
            "la_name":   str(ws.cell_value(i, 3)).strip(),
            "rgn_code":  str(ws.cell_value(i, 4)).strip(),
            "rgn_name":  str(ws.cell_value(i, 5)).strip(),
            "income":    income,
        })
    return records, sheet_is_weekly


def _extract_sheet_xlsx(ws: Any) -> tuple[list[dict], bool]:
    """Parse an openpyxl read-only worksheet. Returns (records, sheet_is_weekly).
    income is the raw value — caller applies ×52 if needed."""
    sheet_is_weekly = "weekly" in ws.title.lower()

    rows = list(ws.iter_rows(values_only=True))
    header_row = None
    for i, row in enumerate(rows[:20]):
        if row[0] and str(row[0]).strip().lower() == "msoa code":
            header_row = i
            break
    if header_row is None:
        return [], sheet_is_weekly

    records = []
    for row in rows[header_row + 1:]:
        if not row[0]:
            continue
        msoa = str(row[0]).strip()
        if not re.match(r"^[EW]\d{8}$", msoa):
            continue
        income = _to_float(row[6])
        if income is None:
            continue
        records.append({
            "msoa_code": msoa,
            "msoa_name": str(row[1]).strip() if row[1] else "",
            "la_code":   str(row[2]).strip() if row[2] else "",
            "la_name":   str(row[3]).strip() if row[3] else "",
            "rgn_code":  str(row[4]).strip() if row[4] else "",
            "rgn_name":  str(row[5]).strip() if row[5] else "",
            "income":    income,
        })
    return records, sheet_is_weekly


def _year_from_workbook_xls(wb: Any) -> int | None:
    """Search first 4 sheets of an xlrd workbook for a financial year reference."""
    for sname in wb.sheet_names()[:5]:
        ws = wb.sheet_by_name(sname)
        for i in range(min(15, ws.nrows)):
            text = " ".join(str(ws.cell_value(i, j)) for j in range(ws.ncols) if ws.cell_value(i, j))
            yr = _year_from_text(text)
            if yr:
                return yr
    return None


def _year_from_workbook_xlsx(wb: Any) -> int | None:
    """Search first 4 sheets of an openpyxl workbook for a financial year reference."""
    for sname in wb.sheetnames[:5]:
        ws = wb[sname]
        for row in list(ws.iter_rows(values_only=True))[:15]:
            text = " ".join(str(c) for c in row if c)
            yr = _year_from_text(text)
            if yr:
                return yr
    return None


def _parse_file(path: Path, crosswalk: dict[str, str]) -> list[dict[str, Any]]:
    """Parse one ONS income file; returns list of dicts with financial_year added."""
    logger.info("  Parsing: {}", path.name)

    if path.suffix.lower() in {".xlsx"}:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        fy_ending = _year_from_workbook_xlsx(wb)
        # File-level: if only weekly income sheets exist, all data in this file is weekly
        file_is_weekly = "Total weekly income" in wb.sheetnames and "Total annual income" not in wb.sheetnames
        # Pick best sheet
        chosen_sheet = None
        for preferred in PREFERRED_SHEETS:
            if preferred in wb.sheetnames:
                chosen_sheet = preferred
                break
        if chosen_sheet is None:
            logger.warning("    No usable sheet found in {}; skipping", path.name)
            wb.close()
            return []
        records, sheet_weekly = _extract_sheet_xlsx(wb[chosen_sheet])
        is_weekly = sheet_weekly or file_is_weekly
        wb.close()

    else:  # .xls
        import xlrd
        wb = xlrd.open_workbook(path)
        fy_ending = _year_from_workbook_xls(wb)
        # File-level: if only weekly income sheets exist, all data in this file is weekly
        file_is_weekly = "Total weekly income" in wb.sheet_names() and "Total annual income" not in wb.sheet_names()
        chosen_sheet = None
        for preferred in PREFERRED_SHEETS:
            if preferred in wb.sheet_names():
                chosen_sheet = preferred
                break
        if chosen_sheet is None:
            logger.warning("    No usable sheet found in {}; skipping", path.name)
            return []
        records, sheet_weekly = _extract_sheet_xls(wb.sheet_by_name(chosen_sheet))
        is_weekly = sheet_weekly or file_is_weekly

    if fy_ending is None:
        logger.warning("    Could not detect year from {}; skipping", path.name)
        return []

    # Convert FY-ending year to start-year convention
    # "financial year ending 2020" (Apr 2019 – Mar 2020) → stored as 2019
    financial_year = fy_ending - 1

    logger.info("    Sheet='{}' | FY ending {} → stored year {} | is_weekly={} | {} MSOAs",
                chosen_sheet, fy_ending, financial_year, is_weekly, len(records))

    if not records:
        logger.warning("    No MSOA rows extracted from {}", path.name)
        return []

    # Apply weekly→annual conversion here (raw values not yet multiplied)
    if is_weekly:
        for r in records:
            r["income"] = r["income"] * 52

    # Sanity-check: median income should be between £5k and £150k
    sample_incomes = sorted(r["income"] for r in records)[len(records) // 2 - 1]
    if not (5_000 < sample_incomes < 150_000):
        logger.warning("    Median income sample={:,.0f} looks implausible — check weekly/annual flag",
                       sample_incomes)

    # Apply MSOA 2011→2021 crosswalk (only affects 2011-boundary files)
    remapped = 0
    for r in records:
        new_code = crosswalk.get(r["msoa_code"])
        if new_code:
            r["msoa_code"] = new_code
            remapped += 1
    if remapped:
        logger.info("    Remapped {} MSOA codes to 2021 boundary", remapped)

    for r in records:
        r["financial_year"] = financial_year

    return records


# ─── Main ───────────────────────────────────────────────────────────────────

@click.command()
@click.option("--income-dir", default=None, type=click.Path(path_type=Path),
              help="Override default data/raw/income directory")
def main(income_dir: Path | None) -> None:
    src_dir = income_dir or INCOME_DIR
    out     = DATA_STAGED / "stg_income.parquet"

    files = sorted(
        p for p in src_dir.iterdir()
        if p.suffix.lower() in {".xlsx", ".xls"} and p.stem.lower() != "readme"
    )
    if not files:
        raise FileNotFoundError(
            f"No .xlsx/.xls files found in {src_dir}\n"
            "Download ONS income files — see data/raw/income/README.md"
        )
    logger.info("Found {} income file(s)", len(files))

    crosswalk = _load_crosswalk()

    all_rows: list[dict[str, Any]] = []
    for path in files:
        rows = _parse_file(path, crosswalk)
        all_rows.extend(rows)

    if not all_rows:
        raise ValueError("No income data parsed from any file.")

    # ── Load into DuckDB for dedup + enrichment ──────────────────────────────
    con = duckdb.connect()
    con.execute(f"PRAGMA threads={DUCKDB_THREADS}")
    con.execute(f"PRAGMA memory_limit='{DUCKDB_MEMORY}'")

    con.execute("""
        CREATE TABLE raw_income (
            msoa_code       VARCHAR,
            msoa_name       VARCHAR,
            la_code         VARCHAR,
            la_name         VARCHAR,
            rgn_code        VARCHAR,
            rgn_name        VARCHAR,
            income          DOUBLE,
            financial_year  INTEGER
        )
    """)

    batch_size = 10_000
    for i in range(0, len(all_rows), batch_size):
        batch = all_rows[i : i + batch_size]
        con.executemany(
            "INSERT INTO raw_income VALUES (?,?,?,?,?,?,?,?)",
            [(r["msoa_code"], r["msoa_name"], r["la_code"], r["la_name"],
              r["rgn_code"], r["rgn_name"], r["income"], r["financial_year"])
             for r in batch],
        )

    total_in = con.execute("SELECT COUNT(*) FROM raw_income").fetchone()[0]
    logger.info("Inserted {:,} raw rows; deduplicating…", total_in)

    # Deduplicate: if same MSOA × year appears twice (unlikely but safe), keep
    # the one with higher income (avoids accidentally keeping weekly × 52 twice).
    con.execute("""
        CREATE TABLE raw_dedup AS
        SELECT *
        FROM (
            SELECT *,
                ROW_NUMBER() OVER (
                    PARTITION BY msoa_code, financial_year
                    ORDER BY income DESC
                ) AS rn
            FROM raw_income
        ) WHERE rn = 1
    """)
    dedup_count = con.execute("SELECT COUNT(*) FROM raw_dedup").fetchone()[0]
    logger.info("After dedup: {:,} rows", dedup_count)

    # Enrich with rgn/ladcd from stg_ons (more reliable than per-file columns)
    if ONS_PARQUET.exists():
        logger.info("Enriching with rgn/ladcd from stg_ons…")
        con.execute(f"""
            CREATE TABLE stg_income AS
            SELECT
                d.msoa_code,
                d.msoa_name,
                COALESCE(o.ladcd,   d.la_code)  AS ladcd,
                COALESCE(o.rgn,     d.rgn_code) AS rgn,
                d.income                        AS net_income_mean,
                d.income                        AS net_income_median,   -- single-value series
                d.financial_year
            FROM raw_dedup d
            LEFT JOIN (
                SELECT msoa21, ANY_VALUE(ladcd) AS ladcd, ANY_VALUE(rgn) AS rgn
                FROM read_parquet('{ONS_PARQUET}')
                WHERE msoa21 IS NOT NULL
                GROUP BY msoa21
            ) o ON d.msoa_code = o.msoa21
            WHERE d.income IS NOT NULL
        """)
    else:
        logger.warning("stg_ons.parquet not found — rgn enrichment skipped")
        con.execute("""
            CREATE TABLE stg_income AS
            SELECT
                msoa_code, msoa_name, la_code AS ladcd, rgn_code AS rgn,
                income AS net_income_mean, income AS net_income_median,
                financial_year
            FROM raw_dedup
            WHERE income IS NOT NULL
        """)

    DATA_STAGED.mkdir(parents=True, exist_ok=True)
    con.execute(f"""
        COPY stg_income TO '{out}'
        (FORMAT PARQUET, CODEC 'ZSTD', COMPRESSION_LEVEL 6)
    """)

    # Summary
    final_n    = con.execute(f"SELECT COUNT(*) FROM '{out}'").fetchone()[0]
    yr_range   = con.execute(f"SELECT MIN(financial_year), MAX(financial_year) FROM '{out}'").fetchone()
    yr_counts  = con.execute(f"""
        SELECT financial_year, COUNT(*) AS n
        FROM '{out}' GROUP BY 1 ORDER BY 1
    """).fetchall()
    rgn_null   = con.execute(f"SELECT COUNT(*) FROM '{out}' WHERE rgn IS NULL").fetchone()[0]

    logger.success("stg_income.parquet written — {:,} rows, FY {}-{}", final_n, yr_range[0], yr_range[1])
    logger.info("Rows per year: {}", {r[0]: r[1] for r in yr_counts})
    if rgn_null:
        logger.warning("{:,} rows have NULL rgn (MSOA not found in stg_ons)", rgn_null)


if __name__ == "__main__":
    main()
